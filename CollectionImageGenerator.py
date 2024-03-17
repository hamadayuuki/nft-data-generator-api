# $ pip install -r requirments.txt

import openpyxl    # Excel処理を簡単に行えるので採用
import requests
from PIL import Image
from datetime import datetime

from io import BytesIO
import json
import os



# =========== Utils ===========
def showSheet(fileName):
    book = openpyxl.load_workbook(fileName)
    sheet = book.active
    for row in sheet.iter_rows(min_row = 1):
        print(f"[ {row[0].row}行目 ]")
        # 行からセルを1個ずつ取得し、処理をする
        for cell in row:
            print(f"{cell.column}列目：{cell.value}")
        print('------------------------------------------')
    book.close()


def getSheetArray(fileName):
    book = openpyxl.load_workbook(fileName)
    sheet = book.active
    saveSheetArray = []    # 最終的にタイトル以外のエクセルデータが全て格納される
    for row in sheet.iter_rows(min_row = 2):    # min_row = 2 : タイトル除外
        saveRow = [cell.value for cell in row if cell.value != None]
        saveSheetArray.append(saveRow)
    book.close()
    sheetArray = [array for array in saveSheetArray if not len(array) == 0]   # remove empty array
    return sheetArray


# 4 item image -> collectionImage
def generateCollectionImage(sampleImagesURL, collectionIconURL, collectionName, outputFolderPath, size = 1000):
    # each item of collection
    width = size // 2
    height = size // 2

    # fetch 4 item image
    images = []   # [Image]
    for i in range(4):
        if i < len(sampleImagesURL): url = sampleImagesURL[i]
        else: url = collectionIconURL    # sampleImageが足りない時は collection icon で補完
        response = requests.get(url)
        img = Image.open(BytesIO(response.content))
        resizedImg = img.resize((width, height))
        images.append(resizedImg)
        resizedImg.save(os.path.join(outputFolderPath, f'{i}.png'))

    # synthesis item image for create collectionImage
    collectionImage = Image.new('RGB', (size, size))   # init
    collectionImage.paste(images[0], (0, 0))           # ↖︎
    collectionImage.paste(images[2], (0, height))      # ↙︎
    collectionImage.paste(images[1], (width, 0))       # ↗︎
    collectionImage.paste(images[3], (width, height))  # ↘︎

    # save collectionImage
    collectionImage.save(os.path.join(outputFolderPath, f'{collectionName}.png'))



# =========== main =============

# settings of generate folder
reporter = "Hamada"   # use folder name
now = datetime.now()
generateTime = now.strftime("%Y-%m%d-%H%M%S")
if not os.path.exists(f"{generateTime}-{reporter}"):
    os.makedirs(f"{generateTime}-{reporter}")

# get collection info from Exel
fileName = "collections.xlsx"
sheetArray = getSheetArray(fileName)

# generate CollectionImage
for collectionName, contractAddress in sheetArray:
    print(f"============ {collectionName} ============")
    outputFolderPath =  os.path.join(f"{generateTime}-{reporter}", collectionName)
    if not os.path.exists(outputFolderPath):
        os.makedirs(outputFolderPath)
    
    url = f'https://api.reservoir.tools/collections/v7?id={contractAddress}'
    headers = {
        'Accept': '*/*',
        'x-api-key': 'demo-api-key',
    }

    resJSON= requests.get(url, headers = headers).json()
    sampleImagesURL = resJSON["collections"][0]["sampleImages"]
    collectionIconURL = resJSON["collections"][0]["image"]
    print("Complete: fetch collection")

    generateCollectionImage(sampleImagesURL, collectionIconURL, collectionName, outputFolderPath)
    print("Complete: generate CollectionImage\n\n")
